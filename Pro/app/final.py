from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import JSONResponse
import shutil
import os
os.environ["TF_ENABLE_ONEDNN_OPTS"] = "0"
# import pickle
# from sklearn.neighbors import KNeighborsClassifier
import cv2
import tensorflow as tf
# from tensorflow.keras.applications import ResNet50
# from tensorflow.keras.applications.resnet50 import preprocess_input
import numpy as np
import pandas as pd
import time
from datetime import datetime
import random
from fastapi.responses import StreamingResponse, JSONResponse
from starlette.staticfiles import StaticFiles
import subprocess
from PIL import Image

from openpyxl import load_workbook


app = FastAPI()

cap = cv2.VideoCapture(0)

# Parent Directory of dataset
# data_upload = "D:/SREC/Dataset_5"


person_in = {"name":""}
person_out = {"name":""}
dept_name = {"dept":""}

dept = dept_name["dept"]


# Define a function to read in an image file and convert it to a numpy array
def read_image(file_path, size=(256, 256)):
    with Image.open(file_path) as img:
        img = img.resize(size)
        img = img.convert('RGB')
        img_data = np.asarray(img)
    return img_data


def normalize_image(img):
    # Convert the image to float data type
    img = img.astype('float32')

    # Normalize the image pixels to have zero mean and unit variance
    img -= np.mean(img)
    img /= np.std(img)

    return img

# Function to preprocess test image
def preprocess_image(image_path):
    img = Image.open(image_path)
    img = img.resize((256, 256))
    img = img.convert('RGB')
    img = np.asarray(img)
    img = img.astype('float32')
    img -= np.mean(img)
    img /= np.std(img)
    img = img.reshape(1, 256, 256, 3)
    return img


upload = "uploads"
isdir = os.path.isdir(upload)
if isdir:
    pass
else:
    os.makedirs(upload)


def get_names(d):
    path = r"D:\SREC\Dataset_5" + "\\" + d
    print("Current Directory :",path)
    return os.listdir(path)

## Entry and Exit
@app.post("/in_out")
async def test_user(check: str, d: str, file: UploadFile = File(...)):
    # if action in ["entry" or "exit"]:
        file_path = os.path.join(upload, file.filename)
        with open(file_path, "wb") as buffer:
                    shutil.copyfileobj(file.file, buffer)
        
        # # Save uploaded file
        # with open(file_path, "wb") as buffer:
        #     buffer.write(file.file.read())

        test_image = preprocess_image(file_path)

        m_path = r"D:\SREC\Model" + "\\" + d + "\\" + "Model.h5"
        loaded_model = tf.keras.models.load_model(m_path)

        class_probabilities = loaded_model.predict(test_image)

     
        predicted_class_index = np.argmax(class_probabilities)
        
        class_names = get_names(d)
        print(class_names)
        person_name = class_names[predicted_class_index]
        if check == "in":
            person_in["name"] = person_name
        if check == "out":
            person_out["name"] = person_name
        dept_name["dept"] = d
        return person_name, d



@app.get("/in")
async def entry_in_2():

    c_t = time.ctime()
    entry_in_2(dept_name["dept"], person_in["name"])
    return {"Name": person_in["name"], "Status": "Checked In!", "Time": c_t}

@app.get("/out")
async def get_out_2():
    c_t = time.ctime()
    get_out_2(dept_name["dept"], person_out["name"])
    return {"Name": person_out["name"], "Status": "Checked Out!", "Time": c_t}
    
def augment_image(image):
    """Applies different augmentation techniques and returns a list of transformed images."""
    augmented_images = []

    # 1. Original image
    augmented_images.append(image)

    # 2. Horizontally flipped
    flipped = cv2.flip(image, 1)
    augmented_images.append(flipped)

    # 3. Rotated by a random angle (-20 to 20 degrees)
    angle = random.randint(-20, 20)
    (h, w) = image.shape[:2]
    M = cv2.getRotationMatrix2D((w//2, h//2), angle, 1.0)
    rotated = cv2.warpAffine(image, M, (w, h))
    augmented_images.append(rotated)

    # 4. Brightness adjustment (increase or decrease)
    brightness_factor = random.uniform(0.7, 1.3)  # Random brightness scale
    brightened = np.clip(image * brightness_factor, 0, 255).astype(np.uint8)
    augmented_images.append(brightened)

    # 5. Gaussian Blur
    blurred = cv2.GaussianBlur(image, (5, 5), 0)
    augmented_images.append(blurred)

    # 6. Scaling (Zoom in by cropping and resizing)
    crop_ratio = 0.9  # Zoom in by keeping 90% of the original size
    crop_h, crop_w = int(h * crop_ratio), int(w * crop_ratio)
    cropped = image[(h - crop_h) // 2: (h + crop_h) // 2, (w - crop_w) // 2: (w + crop_w) // 2]
    zoomed = cv2.resize(cropped, (w, h))  # Resize back to original size
    augmented_images.append(zoomed)

    return augmented_images


@app.post("/live")
async def live_feed(new_user: str):
    cap = cv2.VideoCapture(0)  # Open webcam

    if not cap.isOpened():
        print("Error: Could not open webcam.")
        exit()
        
    c = 0  

    print("Press 'c' to capture an image, 'q' to quit.")

    while True:
        ret, frame = cap.read()
        if not ret:
            print("Failed to grab frame.")
            break

        # Display capture count on screen
        cv2.putText(frame, f"Captures: {c}/5", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 
                    1, (0, 255, 0), 2, cv2.LINE_AA)

        # Show the live video feed
        cv2.imshow("Webcam - Press 'c' to capture, 'q' to quit", frame)

        # Capture key press
        key = cv2.waitKey(1) & 0xFF

        if key == ord('c') and c < 5:  # If 'c' is pressed, capture image
            c += 1
            new_user_folder = os.path.join(data_upload, new_user)
            os.makedirs(new_user_folder, exist_ok=True)  # Create user folder

            # Generate filename for original image (replace spaces with underscores)
            base_filename = f"{new_user.replace(' ', '_')}_{c}"
            filepath = os.path.join(new_user_folder, base_filename + ".jpg")

            # Save the original image
            cv2.imwrite(filepath, frame)
            print(f"Image saved: {base_filename}.jpg")

            # Apply data augmentation
            augmented_images = augment_image(frame)

            # Save augmented images with _aug_{i} suffix
            for i, aug_img in enumerate(augmented_images[1:], start=1):  # Skip original
                aug_filepath = os.path.join(new_user_folder, f"{base_filename}_aug_{i}.jpg")
                cv2.imwrite(aug_filepath, aug_img)
                print(f"Augmented Image saved: {base_filename}_aug_{i}.jpg")

        elif key == ord('q'):  # Quit program
            print("Exiting...")
            break

    # Release resources
    cap.release()
    cv2.destroyAllWindows()

    df = pd.read_excel("D:/SREC/T4.xlsx")
    c_t = time.ctime()
    new_entry = {"Name": new_user, "In": c_t, "Out": "", "Work": "", "Attendance": "", "OT": ""}
    new_df = pd.DataFrame([new_entry]) 
    df = pd.concat([df, new_df], ignore_index=True)
    
    df.to_excel("D:/SREC/T4.xlsx", index=False)
    
    retrain_model()

    return {"Name": new_user, "Checked-In": c_t, "message": "Face captured successfully"}

def retrain_model(d1, d2):

    print("Model is retraiing")
    script_path = r"D:\SREC\Cnn_retrain.py"  # Update with actual script path
    python_path = r"D:\SREC\myenv\Scripts\python.exe"  # Update with actual path

    # subprocess.run([python_path, script_path], check=True)

    subprocess.run([python_path, script_path, d1, d2], check=True)


## To calculate Work
def work(name, dep):
    
    print("Calculate work")
    df = pd.read_excel("D:/SREC/T10.xlsx",sheet_name=dep)
    in_time = df.loc[df["Name"] == name, "In"].values
    out_time = df.loc[df["Name"] == name, "Out"].values
    
    in_list = [x.strip() for x in str(in_time[0]).split(",") if x.strip()]
    out_list = [x.strip() for x in str(out_time[0]).split(",") if x.strip()]

    # Pair timestamps index-wise
    matched_pairs = list(zip(in_list, out_list))
#     w = 0
    
#     # print(matched_pairs[0][0])
#     for i in range(len(matched_pairs)):
#         in_time = matched_pairs[i][0]
#         out_time = matched_pairs[i][1]
#         format_str = "%a %b %d %H:%M:%S %Y"  
#         dt1 = datetime.strptime(in_time, format_str)
#         dt2 = datetime.strptime(out_time, format_str)

# # Calculate the difference in hours
#         time_difference = dt2 - dt1
#         c_w = time_difference.total_seconds() / 3600
#         c_w = "{:.3f}".format(w)
#         c_w = float(c_w)
#         w += c_w
        
     # Ensure both lists have the same length
    min_length = min(len(in_list), len(out_list))
    matched_pairs = list(zip(in_list[:min_length], out_list[:min_length]))

    format_str = "%a %b %d %H:%M:%S %Y"  # Date format
    w = 0 # Total work time in hours
    ot = 0 # Total OT in hours
    c_w = 0
    o_t = 0
    o_t = df.loc[df["Name"]==person_out["name"], "OT"].values
    dt1 = dt2 =None
    for in_time, out_time in matched_pairs:
        dt1 = datetime.strptime(in_time, format_str)
        dt2 = datetime.strptime(out_time, format_str)

        # Calculate the difference in hours
        time_difference = dt2 - dt1
        c_w = time_difference.total_seconds() / 3600  

        # Format to 3 decimal places
        c_w = float("{:.3f}".format(c_w))
        w += c_w  # Accumulate work time
        
    if float(w) >= 10 and dt1.time()<=10:
            ot = float(w)-9
            ot += o_t
    if c_w >= 9:
        attendance = "Present"
    else:
        attendance = "Absent"
    return w, c_w, attendance, ot



@app.post("/live_2")
async def live_feed(new_user: str, dep:str):
  
    new_data_upload = r"D:\SREC\Dataset_5"

    cap = cv2.VideoCapture(0)  # Open webcam
    new_user_dept = "D:\SREC\Dataset_5" + "\\" + dep
    if not os.path.exists(new_user_dept):
     os.makedirs(new_user_dept)
    model_dir = "D:\SREC\Model" + "\\" + dep
    if not os.path.exists(model_dir):
        os.makedirs(model_dir)
    if not cap.isOpened():
        print("Error: Could not open webcam.")
        exit()
        
    c = 0  

    print("Press 'c' to capture an image, 'q' to quit.")

    while True:
        ret, frame = cap.read()
        if not ret:
            print("Failed to grab frame.")
            break

        # Display capture count on screen
        cv2.putText(frame, f"Captures: {c}/5", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 
                    1, (0, 255, 0), 2, cv2.LINE_AA)

        # Show the live video feed
        cv2.imshow("Webcam - Press 'c' to capture, 'q' to quit", frame)

        # Capture key press
        key = cv2.waitKey(1) & 0xFF

        if key == ord('c') and c < 5:  # If 'c' is pressed, capture image
            c += 1
            new_user_folder = os.path.join(new_user_dept, new_user)
            os.makedirs(new_user_folder, exist_ok=True)  # Create user folder

            # Generate filename for original image (replace spaces with underscores)
            base_filename = f"{new_user.replace(' ', '_')}_{c}"
            filepath = os.path.join(new_user_folder, base_filename + ".jpg")

            # Save the original image
            cv2.imwrite(filepath, frame)
            print(f"Image saved: {base_filename}.jpg")

            # Apply data augmentation
            augmented_images = augment_image(frame)

            # Save augmented images with _aug_{i} suffix
            for i, aug_img in enumerate(augmented_images[1:], start=1):  # Skip original
                aug_filepath = os.path.join(new_user_folder, f"{base_filename}_aug_{i}.jpg")
                cv2.imwrite(aug_filepath, aug_img)
                print(f"Augmented Image saved: {base_filename}_aug_{i}.jpg")

        elif key == ord('q'):  # Quit program
            print("Exiting...")
            break

    # Release resources
    cap.release()
    cv2.destroyAllWindows()
    add_new_user(new_user, dep)
    print("Model Retraining")
    retrain_model(new_user_dept, model_dir)
    print("Success")
    # return {"Name": new_user, "Checked-In": c_t, "message": "Face captured successfully"}
    print(f"Entry added for {new_user} in {dep} sheet.")
    return {"Name": new_user, "Checked-In": time.ctime(), "message": "Face captured successfully"}


def get_out_2(dep, name):
    file_path = r"D:/SREC/T10.xlsx"
    df = pd.read_excel(file_path)
    c_t = time.ctime()
    sheet_name = dep
    person_out["name"] = name
    book = load_workbook(file_path)

    if sheet_name in book.sheetnames:
        # Read the specific sheet
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    else:
        # If sheet doesn't exist, create an empty DataFrame
        df = pd.DataFrame(columns=["Name", "In", "Out", "Work", "Attendance", "OT"])  

    c_t = time.ctime()

    v1 = person_out["name"]
    c1 = "Name"
    u1 = "In"  # Column to update
    out_time = c_t

    v2 = person_out["name"]
    c2 = "Name"
    u2 = "Out"
    u3 = "Work"
    n2 = c_t
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    old = df.loc[df["Name"] == person_out["name"], "Out"].notna().any()
    print(person_out["name"])
    print("Old", old)
    if old:
        df.loc[df[c2] == v2, u2] = df.loc[df[c2] == v2, u2].fillna('').apply(lambda x: n2 if x == '' else x + ", " + n2)
        work_hours, c_w, attendance, ot = work(person_out["name"], dep)
        # To write in excel
        df.loc[df["Name"] == person_out["name"], "Work"] = work_hours
        df.loc[df[c2] == v2, "Attendance"] = attendance
        df.loc[df[c2] == v2, "OT"] = ot
    else:
        df.loc[df["Name"] == person_out["name"], "Out"] = c_t
        in_time = df.loc[df[c1] == v1, u1].values[0]
    
        format_str = "%a %b %d %H:%M:%S %Y"  # Example: 'Mon Feb 26 14:30:15 2024'
        dt1 = datetime.strptime(in_time, format_str)
        dt2 = datetime.strptime(out_time, format_str)

        # Calculate the difference in hours
        time_difference = dt2 - dt1
        w = time_difference.total_seconds() / 3600
        c_w = "{:.3f}".format(w)
        
        df.loc[df[c2] == v2, u3] = c_w
        
        if float(c_w) >= 9:
            df.loc[df[c2] == v2, "Attendance"] = "Present"
        else:
            df.loc[df[c2] == v2, "Attendance"] = "Absent"
            
        if float(c_w) >= 10 and dt1.time():
            ot = float(w)-9
            df.loc[df[c2] == v2, "OT"] = ot
        else:
            df.loc[df[c2] == v2, "OT"] = 0

    # FIX: Use ExcelWriter to preserve other sheets
    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
def entry_in_2(dep , name):
    file_path = r"D:/SREC/T10.xlsx"
    c_t = time.ctime()
    sheet_name = dep
    person_in["name"] = name
    # Check if file exists first
    try:
        book = load_workbook(file_path)
        file_exists = True
    except FileNotFoundError:
        file_exists = False
        
    if file_exists and sheet_name in book.sheetnames:
        # Read the specific sheet
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    else:
        # If sheet doesn't exist, create an empty DataFrame
        df = pd.DataFrame(columns=["Name", "In", "Out", "Work", "Attendance", "OT"])
    
    name_col = "Name"
    if name_col in df.columns:
        # Check if the person already exists
        if person_in["name"] in df["Name"].values:
            df.loc[df["Name"] == person_in["name"], "In"] = df.loc[df["Name"] == person_in["name"], "In"].fillna('').astype(str) + ", " + c_t
            print("Success")
        else:
            # Add a new entry
            new_entry = {"Name": person_in["name"], "In": c_t, "Out": "", "Work": "", "Attendance": "", "OT": ""}
            new_df = pd.DataFrame([new_entry])
            df = pd.concat([df, new_df], ignore_index=True)
    else:
        # If the sheet was empty or missing columns, create a new entry
        new_entry = {"Name": person_in["name"], "In": c_t, "Out": "", "Work": "", "Attendance": "", "OT": ""}
        df = pd.DataFrame([new_entry])
    
    # Save back to the Excel file with proper handling of existing sheets
    if file_exists:
        # Get all existing sheets
        all_sheets = book.sheetnames
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            # Write the updated sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Make sure we don't lose other sheets
            if sheet_name not in all_sheets:
                book.create_sheet(sheet_name)
                
    else:
        # Create a new Excel file if it doesn't exist
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

def add_new_user(new_user, dep):
    file_path = r"D:/SREC/T10.xlsx"
    
    # Check if file exists first
    try:
        book = load_workbook(file_path)
        file_exists = True
    except FileNotFoundError:
        book = None
        file_exists = False
    
    sheet_name = dep
    
    # Read the department sheet if it exists
    if file_exists and book is not None and sheet_name in book.sheetnames:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    else:
        # Create empty DataFrame if sheet doesn't exist
        df = pd.DataFrame(columns=["Name", "In", "Out", "Work", "Attendance", "OT"])

    # Add new entry
    c_t = time.ctime()
    new_entry = {"Name": new_user, "In": c_t, "Out": "", "Work": "", "Attendance": "", "OT": ""}
    df = pd.concat([df, pd.DataFrame([new_entry])], ignore_index=True)

    # Save back to the Excel file with proper handling of existing sheets
    if file_exists and book is not None:
        # Get all existing sheets
        all_sheets = book.sheetnames
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            # Write the updated sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Make sure we don't lose other sheets
            if sheet_name not in all_sheets:
                book.create_sheet(sheet_name)
    else:
        # Create a new Excel file if it doesn't exist
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    print(f"Entry added for {new_user} in {dep} sheet.")
    