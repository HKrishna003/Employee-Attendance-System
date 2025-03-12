from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from fastapi.responses import JSONResponse
import shutil
import os
os.environ["TF_ENABLE_ONEDNN_OPTS"] = "0"
# import pickle
# from sklearn.neighbors import KNeighborsClassifier
import cv2
import tensorflow as tf
# from tensorflow.keras.applications import ResNet50
# from tensorflow.keras.applications.resnet50 import preprocess_input
import numpy as np
import pandas as pd
import time
from datetime import datetime
import random
from fastapi.responses import StreamingResponse, JSONResponse
from starlette.staticfiles import StaticFiles
import subprocess
from PIL import Image
from typing import List
from openpyxl import load_workbook


app = FastAPI()

cap = cv2.VideoCapture(0)

# Parent Directory of dataset
# data_upload = "D:/SREC/Dataset_5"


person_in = {"name":""}
person_out = {"name":""}
dept_name = {"dept":""}

dept = dept_name["dept"]


# Define a function to read in an image file and convert it to a numpy array
def read_image(file_path, size=(256, 256)):
    with Image.open(file_path) as img:
        img = img.resize(size)
        img = img.convert('RGB')
        img_data = np.asarray(img)
    return img_data


def normalize_image(img):
    # Convert the image to float data type
    img = img.astype('float32')

    # Normalize the image pixels to have zero mean and unit variance
    img -= np.mean(img)
    img /= np.std(img)

    return img

# Function to preprocess test image
def preprocess_image(image_path):
    img = Image.open(image_path)
    img = img.resize((256, 256))
    img = img.convert('RGB')
    img = np.asarray(img)
    img = img.astype('float32')
    img -= np.mean(img)
    img /= np.std(img)
    img = img.reshape(1, 256, 256, 3)
    return img


upload = "uploads"
isdir = os.path.isdir(upload)
if isdir:
    pass
else:
    os.makedirs(upload)


def get_names(d):
    path =r"D:\srec_project\project\backend_1\Dataset_5" + "\\" + d
    print("Current Directory :",path)
    return os.listdir(path)

u = 0

UPLOAD_DIR = "captured_images"  # Directory where images will be saved
os.makedirs(UPLOAD_DIR, exist_ok=True)  # Ensure the directory exists

# @app.post("/in_out")
# async def test_user(check: str, d: str):
#     global u  # Ensure we modify the global counter

#     cap = cv2.VideoCapture(0)
#     if not cap.isOpened():
#         return {"Error": "Could not open webcam"}

#     while True:
#         ret, frame = cap.read()
#         if not ret:
#             print("Failed to grab frame.")
#             break

#         # Display capture count on screen
#         cv2.putText(frame, f"Captures: {u}", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 
#                     1, (0, 255, 0), 2, cv2.LINE_AA)

#         # Show the live video feed
#         cv2.imshow("Webcam - Press 'c' to capture, 'q' to quit", frame)

#         key = cv2.waitKey(1) & 0xFF
#         if key == ord('c'):  # Capture image when 'c' is pressed
#             file_path = os.path.join(UPLOAD_DIR, f"u_{u}.jpg")  # Correct path
#             cv2.imwrite(file_path, frame)
#             print(f"Image saved at: {file_path}")
#             break  # Exit loop after capturing an image

#     cap.release()
#     cv2.destroyAllWindows()

#     # Process the image
#     test_image = preprocess_image(file_path)  # Ensure preprocess_image is defined

#     # Load model dynamically based on department
#     model_path = os.path.join(r"D:\srec_project\project\backend_1\Model", d, "Model.h5")
#     if not os.path.exists(model_path):
#         return {"Error": f"Model file not found at {model_path}"}

#     loaded_model = tf.keras.models.load_model(model_path)
#     class_probabilities = loaded_model.predict(test_image)
#     predicted_class_index = np.argmax(class_probabilities)

#     class_names = get_names(d)  # Ensure this function is defined
#     person_name = class_names[predicted_class_index]

#     # Store entry/exit info
#     person_name = class_names[predicted_class_index]
#     if check == "in":
#         person_in["name"] = person_name
#     if check == "out":
#         person_out["name"] = person_name
#     dept_name["dept"] = d
#         # return person_name, d


#     # dept_name = {"dept": d}
#     u += 1  # Increment counter

#     return person_name, d



# @app.get("/in")
# async def entry_in_2():
#     c_t = time.ctime()
#     entry_in_2(dept_name["dept"], person_in["name"])
#     return {"Name": person_in["name"], "Status": "Checked In!", "Time": c_t}

# @app.get("/out")
# async def get_out():
#     c_t = time.ctime()
#     get_out_2(dept_name["dept"], person_out["name"])
#     return {"Name": person_out["name"], "Status": "Checked Out!", "Time": c_t}



@app.post("/in_out")
async def process_attendance(
    check: str = Form(...),
    d: str = Form(...),
    image: UploadFile = File(...)
):
    global u  # Ensure we modify the global counter

    # Create upload directory if it doesn't exist
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    
    # Save the uploaded image
    file_path = os.path.join(UPLOAD_DIR, f"u_{u}.jpg")
    
    # Read and save the uploaded image
    contents = await image.read()
    with open(file_path, "wb") as f:
        f.write(contents)
    
    print(f"Image saved at: {file_path}")
    
    # Process the image
    try:
        test_image = preprocess_image(file_path)  # Ensure preprocess_image is defined
        
        # Load model dynamically based on department
        model_path = os.path.join(r"D:\srec_project\project\backend_1\Model", d, "Model.h5")
        if not os.path.exists(model_path):
            return {"Error": f"Model file not found at {model_path}"}
            
        loaded_model = tf.keras.models.load_model(model_path)
        class_probabilities = loaded_model.predict(test_image)
        predicted_class_index = np.argmax(class_probabilities)
            
        class_names = get_names(d)  # Ensure this function is defined
        person_name = class_names[predicted_class_index]
            
        # Store entry/exit info
        if check == "in":
            person_in["name"] = person_name
            dept_name["dept"] = d
            
            # Call the check-in function
            c_t = time.ctime()
            entry_in_2(dept_name["dept"], person_in["name"])
            
            response_data = {
                "status": "success",
                "person_name": person_name,
                "department": d,
                "check_type": check,
                "time": c_t,
                "message": "Checked In!"
            }
            
        elif check == "out":
            person_out["name"] = person_name
            dept_name["dept"] = d
            
            # Call the check-out function
            c_t = time.ctime()
            get_out_2(dept_name["dept"], person_out["name"])
            
            response_data = {
                "status": "success",
                "person_name": person_name,
                "department": d,
                "check_type": check,
                "time": c_t,
                "message": "Checked Out!"
            }
        else:
            return {"Error": "Invalid check type. Must be 'in' or 'out'"}
            
        u += 1  # Increment counter
        return response_data
    
    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }

def retrain_model(d1, d2):

    print("Model is retraiing")
    script_path = r"D:\srec_project\project\backend_1\Cnn_retrain.py"  # Update with actual script path
    python_path = r"D:\srec_project\project\backend_1\myenv\Scripts\python.exe"  # Update with actual path

    # subprocess.run([python_path, script_path], check=True)

    subprocess.run([python_path, script_path, d1, d2], check=True)


## To calculate Work
def work(name, dep):
    
    print("Calculate work")
    df = pd.read_excel("D:/SREC/T10.xlsx",sheet_name=dep)
    in_time = df.loc[df["Name"] == name, "In"].values
    out_time = df.loc[df["Name"] == name, "Out"].values
    
    in_list = [x.strip() for x in str(in_time[0]).split(",") if x.strip()]
    out_list = [x.strip() for x in str(out_time[0]).split(",") if x.strip()]

    # Pair timestamps index-wise
    matched_pairs = list(zip(in_list, out_list))

        
     # Ensure both lists have the same length
    min_length = min(len(in_list), len(out_list))
    matched_pairs = list(zip(in_list[:min_length], out_list[:min_length]))

    format_str = "%a %b %d %H:%M:%S %Y"  # Date format
    w = 0 # Total work time in hours
    ot = 0 # Total OT in hours
    c_w = 0
    o_t = 0
    o_t = df.loc[df["Name"]==person_out["name"], "OT"].values
    dt1 = dt2 =None
    for in_time, out_time in matched_pairs:
        dt1 = datetime.strptime(in_time, format_str)
        dt2 = datetime.strptime(out_time, format_str)

        # Calculate the difference in hours
        time_difference = dt2 - dt1
        c_w = time_difference.total_seconds() / 3600  

        # Format to 3 decimal places
        c_w = float("{:.3f}".format(c_w))
        w += c_w  # Accumulate work time
        
    if float(w) >= 10 and dt1.time()<=10:
            ot = float(w)-9
            ot += o_t
    if c_w >= 9:
        attendance = "Present"
    else:
        attendance = "Absent"
    return w, c_w, attendance, ot


def augment_image(image):
    """Applies different augmentation techniques and returns a list of transformed images."""
    augmented_images = []

    # 1. Original image
    augmented_images.append(image)

    # 2. Horizontally flipped
    flipped = cv2.flip(image, 1)
    augmented_images.append(flipped)

    # 3. Rotated by a random angle (-20 to 20 degrees)
    angle = random.randint(-20, 20)
    (h, w) = image.shape[:2]
    M = cv2.getRotationMatrix2D((w//2, h//2), angle, 1.0)
    rotated = cv2.warpAffine(image, M, (w, h))
    augmented_images.append(rotated)

    # 4. Brightness adjustment (increase or decrease)
    brightness_factor = random.uniform(0.7, 1.3)  # Random brightness scale
    brightened = np.clip(image * brightness_factor, 0, 255).astype(np.uint8)
    augmented_images.append(brightened)

    # 5. Gaussian Blur
    blurred = cv2.GaussianBlur(image, (5, 5), 0)
    augmented_images.append(blurred)

    # 6. Scaling (Zoom in by cropping and resizing)
    crop_ratio = 0.9  # Zoom in by keeping 90% of the original size
    crop_h, crop_w = int(h * crop_ratio), int(w * crop_ratio)
    cropped = image[(h - crop_h) // 2: (h + crop_h) // 2, (w - crop_w) // 2: (w + crop_w) // 2]
    zoomed = cv2.resize(cropped, (w, h))  # Resize back to original size
    augmented_images.append(zoomed)

    return augmented_images


from fastapi.responses import JSONResponse
@app.post("/live_2")
async def live_feed(
    name: str = Form(...),
    department: str = Form(...),
    images: List[UploadFile] = File(...)
):
    try:
        print(f"Received name: {name}, department: {department}, images: {len(images)}")  # Debugging

        # Ensure exactly 30 images are uploaded
        if len(images) != 30:
            return JSONResponse(content={"error": "Exactly 30 images must be uploaded."}, status_code=400)

        # Define the base directory for dataset
        base_dir = r"D:\srec_project\project\backend_1\Dataset_5"

        # Create department directory if it does not exist
        department_dir = os.path.join(base_dir, department)
        os.makedirs(department_dir, exist_ok=True)

        # Create user directory inside the department folder
        user_dir = os.path.join(department_dir, name.replace(" ", "_"))
        os.makedirs(user_dir, exist_ok=True)
        model_dir = r"D:\srec_project\project\backend_1\Model" + "\\" + department
        if not os.path.exists(model_dir):
            os.makedirs(model_dir)
        # Process and save each image
        for idx, image in enumerate(images):
            image_bytes = await image.read()
            nparr = np.frombuffer(image_bytes, np.uint8)
            frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

            if frame is None:
                return JSONResponse(content={"error": f"Failed to decode image {image.filename}"}, status_code=400)

            # Save image inside the user's directory
            filename = f"{idx + 1}.jpg"
            filepath = os.path.join(user_dir, filename)
            cv2.imwrite(filepath, frame)
        add_new_user(name, department)
        print("Model Retraining")
        retrain_model(department_dir,model_dir)
        print("Success")
        return JSONResponse(content={
            "Name": name,
            "Department": department,
            "User Directory": user_dir,
            "Checked-In": time.ctime(),
            "message": "Face captured and stored successfully",
        }, status_code=200)

    except Exception as e:
        print("Error:", str(e))  # Debugging
        return JSONResponse(content={"error": str(e)}, status_code=500)
    
def get_out_2(dep, name):

    file_path = r"D:\srec_project\project\backend_1\T1.xlsx"
    df = pd.read_excel(file_path)
    c_t = time.ctime()
    sheet_name = dep
    person_out["name"] = name
    book = load_workbook(file_path)

    if sheet_name in book.sheetnames:
        # Read the specific sheet
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    else:
        # If sheet doesn't exist, create an empty DataFrame
        df = pd.DataFrame(columns=["Name", "In", "Out", "Work", "Attendance", "OT"])  

    c_t = time.ctime()

    v1 = person_out["name"]
    c1 = "Name"
    u1 = "In"  # Column to update
    out_time = c_t

    v2 = person_out["name"]
    c2 = "Name"
    u2 = "Out"
    u3 = "Work"
    n2 = c_t
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    old = df.loc[df["Name"] == person_out["name"], "Out"].notna().any()
    print(person_out["name"])
    print("Old", old)
    if old:
        df.loc[df[c2] == v2, u2] = df.loc[df[c2] == v2, u2].fillna('').apply(lambda x: n2 if x == '' else x + ", " + n2)
        work_hours, c_w, attendance, ot = work(person_out["name"], dep)
        # To write in excel
        df.loc[df["Name"] == person_out["name"], "Work"] = work_hours
        df.loc[df[c2] == v2, "Attendance"] = attendance
        df.loc[df[c2] == v2, "OT"] = ot
    else:
        df.loc[df["Name"] == person_out["name"], "Out"] = c_t
        in_time = df.loc[df[c1] == v1, u1].values[0]
    
        format_str = "%a %b %d %H:%M:%S %Y"  # Example: 'Mon Feb 26 14:30:15 2024'
        dt1 = datetime.strptime(in_time, format_str)
        dt2 = datetime.strptime(out_time, format_str)

        # Calculate the difference in hours
        time_difference = dt2 - dt1
        w = time_difference.total_seconds() / 3600
        c_w = "{:.3f}".format(w)
        
        df.loc[df[c2] == v2, u3] = c_w
        
        if float(c_w) >= 9:
            df.loc[df[c2] == v2, "Attendance"] = "Present"
        else:
            df.loc[df[c2] == v2, "Attendance"] = "Absent"
            
        if float(c_w) >= 10 and dt1.time():
            ot = float(w)-9
            df.loc[df[c2] == v2, "OT"] = ot
        else:
            df.loc[df[c2] == v2, "OT"] = 0

    # FIX: Use ExcelWriter to preserve other sheets
    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def entry_in_2(dep , name):
    file_path = r"D:\srec_project\project\backend_1\T1.xlsx"
    c_t = time.ctime()
    sheet_name = dep
    person_in["name"] = name
    # Check if file exists first
    try:
        book = load_workbook(file_path)
        file_exists = True
    except FileNotFoundError:
        file_exists = False
        
    if file_exists and sheet_name in book.sheetnames:
        # Read the specific sheet
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    else:
        # If sheet doesn't exist, create an empty DataFrame
        df = pd.DataFrame(columns=["Name", "In", "Out", "Work", "Attendance", "OT"])
    
    name_col = "Name"
    if name_col in df.columns:
        # Check if the person already exists
        if person_in["name"] in df["Name"].values:
            df.loc[df["Name"] == person_in["name"], "In"] = df.loc[df["Name"] == person_in["name"], "In"].fillna('').astype(str) + ", " + c_t
            print("Success")
        else:
            # Add a new entry
            new_entry = {"Name": person_in["name"], "In": c_t, "Out": "", "Work": "", "Attendance": "", "OT": ""}
            new_df = pd.DataFrame([new_entry])
            df = pd.concat([df, new_df], ignore_index=True)
    else:
        # If the sheet was empty or missing columns, create a new entry
        new_entry = {"Name": person_in["name"], "In": c_t, "Out": "", "Work": "", "Attendance": "", "OT": ""}
        df = pd.DataFrame([new_entry])
    
    # Save back to the Excel file with proper handling of existing sheets
    if file_exists:
        # Get all existing sheets
        all_sheets = book.sheetnames
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            # Write the updated sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Make sure we don't lose other sheets
            if sheet_name not in all_sheets:
                book.create_sheet(sheet_name)
                
    else:
        # Create a new Excel file if it doesn't exist
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

def add_new_user(new_user, dep):
    file_path = r"D:\srec_project\project\backend_1\T1.xlsx"
    
    # Check if file exists first
    try:
        book = load_workbook(file_path)
        file_exists = True
    except FileNotFoundError:
        book = None
        file_exists = False
    
    sheet_name = dep
    
    # Read the department sheet if it exists
    if file_exists and book is not None and sheet_name in book.sheetnames:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    else:
        # Create empty DataFrame if sheet doesn't exist
        df = pd.DataFrame(columns=["Name", "In", "Out", "Work", "Attendance", "OT"])

    # Add new entry
    c_t = time.ctime()
    new_entry = {"Name": new_user, "In": c_t, "Out": "", "Work": "", "Attendance": "", "OT": ""}
    df = pd.concat([df, pd.DataFrame([new_entry])], ignore_index=True)

    # Save back to the Excel file with proper handling of existing sheets
    if file_exists and book is not None:
        # Get all existing sheets
        all_sheets = book.sheetnames
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            # Write the updated sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Make sure we don't lose other sheets
            if sheet_name not in all_sheets:
                book.create_sheet(sheet_name)
    else:
        # Create a new Excel file if it doesn't exist
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    print(f"Entry added for {new_user} in {dep} sheet.")
    